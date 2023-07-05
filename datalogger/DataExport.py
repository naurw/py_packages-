import os
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import workbook 
from openpyxl.styles import PatternFill
from time import sleep 
import mimetypes 
from tqdm import tqdm 
import platform 
import subprocess

# os.getcwd()
# import sys 
# sys.path.append(os.path.join(os.getcwd(), 'Desktop/py_packages-/datalogger/gui_add_ons'))
# from pbar import ProgressBar 

workbook.WB_VIEW_NORMAL = 'Sheet View'

class DataExport:
    def __init__(self, dataframe, file_name='v3_speed_test_logs.xlsx', directory='v3_logs'):
        self.df = dataframe
        self.current_path = os.getcwd()
        self.file_name = file_name
        self.directory = directory
        self.new_path = ''
        self.max_col_width = None

        self.check_directory()
        self.write_to_excel()

    def check_directory(self):
        logs_dir = os.path.join(self.current_path, self.directory)
        if not os.path.exists(logs_dir):
            print(f'File Path: {logs_dir} does not exist... creating directory now')

            os.mkdir(logs_dir)
        else:
            print(f'File Path: {logs_dir} exists... changing directory')

        os.chdir(logs_dir)
        self.new_path = os.getcwd()
        print(self.new_path)
    
    def adjust_col_width(self, sheet): 
        if self.max_col_width is None:  
            self.set_max_col_width()

        for col in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = self.get_max_cell_length(sheet[column_letter])
            adjusted_width = self.max_col_width if self.max_col_width is not None and (self.max_col_width >= max_length or self.max_col_width <= max_length + 5) else min(self.max_col_width, max_length +1) 
            ''' 
            Adjusted width is the comparison between the maximum desired width and the maximum string length in the column with the allowance of 5 additonal characters. 
            If the maximum desired width is sufficient (greater than the max_length) it will be used. 
            If the maxmimum desired width is NOT sufficient, the width will be the minimum value between the maximum desired width and ensures that it is at least 1 character winder than the longest string to provide some padding.
            '''
            # adjusted_width = min(self.max_col_width, max_length + 1)  # Adjusted width is the minimum between maximum desired width and the maximum desired lenght of the string
            sheet.column_dimensions[column_letter].width = adjusted_width 

    def set_max_col_width(self): 
        try: 
            self.max_col_width = int(input(f'Enter the maximum desired column width (1-50): '))
            if self.max_col_width > 50 or self.max_col_width < 1: 
                print(f'Invalid input. Using default max value of 50.')
                self.max_col_width = 50

        except ValueError: 
            print('Invalid input. Using default max value of columns.')
            self.max_col_width = self.get_max_cell_length

    def get_max_cell_length(self, column_cells):
        max_length = 0 
        for cell in column_cells: 
            try: 
                cell_value_length = len(str(cell.value))
                if cell_value_length > max_length: 
                    max_length = cell_value_length
                    print(max_length)

            except Exception as e: 
                print(f'Error occured: {e}')

        return max_length 


###############################################################
### Simpler method for computing column widths using pixels ###
###############################################################
# from openpyxl.utils import column_width_from_pixel
#     def adjust_col_width_pixel(self, sheet): 
#         from openpyxl.utils import column_width_from_pixel
#         for col in sheet.columns:
#             max_length = max(len(str(cell.value)) for cell in col)
#             adjusted_width = column_width_from_pixel(max_length * 7)  # assuming default font size of 11pt
#             col[0].column_dimensions.width = adjusted_width

    def write_to_excel(self):
        try:
            file_path = os.path.join(self.new_path, self.file_name)
            if not os.path.exists(file_path):

                if isinstance(self.df, pd.DataFrame) and not isinstance(self.df, list):
                    self.write_single_dataframe(file_path)
                elif isinstance(self.df, list) and all(isinstance(df, pd.DataFrame) for df in self.df):
                    self.write_multiple_dataframes(file_path)
                else:
                    print('Var used is neither a dataframe nor a list of dataframes. Retry.')
                    raise ValueError
            else:

                if isinstance(self.df, pd.DataFrame) and not isinstance(self.df, list):
                    self.append_single_dataframe(file_path)
                elif isinstance(self.df, list) and all(isinstance(df, pd.DataFrame) for df in self.df):
                    self.append_multiple_dataframes(file_path)
                else:
                    print('Var used is neither a dataframe nor a list of dataframes. Retry.')
                    raise ValueError
            os.chdir(os.pardir)
            self.open_file()
        except Exception as e:

            sleep(3)
            if isinstance(self.df, pd.DataFrame) and not isinstance(self.df, list):
                self.append_single_dataframe_alternative()
            elif isinstance(self.df, list) and all(isinstance(df, pd.DataFrame) for df in self.df):
                self.append_multiple_dataframes_alternative()
            else:
                print('Var used is neither a dataframe nor a list of dataframes. Retry.')
                raise ValueError
            os.chdir(os.pardir)
            self.open_file()
    def write_single_dataframe(self, file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            self.df.to_excel(writer, sheet_name='Sheet1', index=False)
            sheet = writer.book['Sheet1']
            self.adjust_col_width(sheet)

    def write_multiple_dataframes(self, file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            for i, df in enumerate(self.df):
                sheet_name = df.name if hasattr(df, 'name') else f'Sheet{i+1}'
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                sheet = writer.book[sheet_name]
                self.adjust_col_width(sheet)

    def append_single_dataframe(self, file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            self.df.to_excel(writer, sheet_name='Sheet1', index=False)
            sheet = writer.book['Sheet1']
            self.adjust_col_width(sheet)

    def append_multiple_dataframes(self, file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            for i, df in enumerate(self.df):
                sheet_name = df.name if hasattr(df, 'name') else f'Sheet{i+1}'
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                sheet = writer.book[sheet_name]
                self.adjust_col_width(sheet)

    def append_single_dataframe_alternative(self):
        with pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a', if_sheet_exists='replace', engine_kwargs={'options': {'sheet_state' : 'visible'}}) as writer:
            try: 
                if mimetypes.guess_type(self.file_name)[0] in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                    print('File is the formats: vnd.ms-excel or vnd.openxml formats')
                    existing_df = pd.read_excel(self.file_name)
                else: 
                    print('File is not the in the accepted pd.read_excel formats... using alternative method') 
                    existing_file = pd.ExcelFile(self.file_name)
                    existing_df = existing_file.parse(sheet_name = 'Sheet1')
            except Exception as e: 
                print(f'Error occured: {e}')

            book = load_workbook(self.file_name)
            writer.book = book
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets} # get all the sheets 

            existing_columns = set() # extract existing column names form the first sheet 
            for sheet in book.worksheets:
                if sheet.title == 'Sheet1':
                    existing_df_sheet = pd.read_excel(self.file_name, sheet_name=sheet.title)
                    existing_columns.update(existing_df_sheet.columns)


            new_columns = set(self.df.columns) # check if new data has the same column names as the existing data 
            if existing_columns == new_columns:
                new_df = pd.concat([existing_df_sheet, self.df], axis=0, ignore_index=True)
                last_row = len(existing_df_sheet)
                new_df.style.apply(lambda x: ['background-color: #C4D9F2' if x.name >= last_row else ''], axis=1) # change color for only newly appended data and default white if data is old 
                new_df.to_excel(writer, sheet_name='Sheet1', index=False)
                sheet = book['Sheet1']
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, values_only=True), start=2): # iterates over the rows of the worksheet starting from a set starting row. returns iterable of Cell objects for each row. Ensures we get the cell values instead of ALL of the cells' objects. Enumerate to get both the row index and the row values during iteration. Start =2 sets the starting value of the row index
                    for col_idx, value in enumerate(row, 1):
                        if row_idx > last_row + 1: 
                            cell = sheet.cell(row = row_idx, column = col_idx)
                            cell.fill = PatternFill(start_color='C4D9F2', end_color='C4D9F2', fill_type='solid')
                        
                self.adjust_col_width(sheet)
            else:
                self.df.to_excel(writer, sheet_name='Sheet1', index=False)
                sheet = book['Sheet1']
                self.adjust_col_width(sheet)

    def append_multiple_dataframes_alternative(self): 
        with pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a', if_sheet_exists='replace', engine_kwargs = {'options' : {'sheet_state' : 'visible'}}) as writer:
            try: 
                if mimetypes.guess_type(self.file_name)[0] in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                    print('File is the formats: vnd.ms-excel or vnd.openxml formats')
                    existing_df = pd.read_excel(self.file_name)
                else: 
                    print('File is not the in the accepted pd.read_excel formats... using alt method') 
                    existing_file = pd.ExcelFile(self.file_name)
                    existing_df = existing_file.parse(sheet_name = 'Sheet1')
            except Exception as e: 
                print(f'Error occured: {e}')

            new_data = []
            existing_columns = set() # Extract column names from the existing data into a set (different from a list because unique values only) 
            for sheet_name, df in existing_df.items():
                existing_columns.update(df.columns)

            for df in self.df:
                new_columns = set(df.columns) # Extract column names from the new dataframes into a set

                if existing_columns.intersection(new_columns): # If there are any overlapping columns, concatenate the new data with the existing data and add to the list of new data to be written
                    existing_df_sheet = existing_df.get(df.name, pd.DataFrame()) # if there is a dataframe name then it takes that. if there isn't then default value 
                    last_row = len(existing_df_sheet)
                    new_df = pd.concat([existing_df_sheet, df], axis=0, ignore_index=True)
                    new_df.style.apply(lambda x: ['background-color: #C4D9F2' if x.name >= last_row else ''], axis=1) # temporary function for creating background colors 
                    new_data.append(new_df)
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, values_only=True), start=2): # iterates over the rows of the worksheet starting from a set starting row. returns iterable of Cell objects for each row. Ensures we get the cell values instead of ALL of the cells' objects. Enumerate to get both the row index and the row values during iteration. Start =2 sets the starting value of the row index
                        for col_idx, value in enumerate(row, 1):
                            if row_idx > last_row + 1: 
                                cell = sheet.cell(row = row_idx, column = col_idx)
                                cell.fill = PatternFill(start_color='C4D9F2', end_color='C4D9F2', fill_type='solid')
                else:
                    new_data.append(df) # If there are no overlapping columns, it treats it as a new dataframe and add it to the list of dataframes 
                
            for df in new_data:
                df.to_excel(writer, sheet_name=df.name, index=False)
                sheet = writer.sheets[df.name]
                self.adjust_col_width(sheet)

    def open_file(self): 
        try: 
            # for i in tqdm(range(100), desc='Opening file', unit='%', ncols=70): # progress bar length set to 70 chars wide 
            #     sleep(0.03) # animation set at 0.03

            with tqdm(total = 100, desc = 'Opening File', unit = '%', ncols = 70) as pbar: 
                for i in range(100): 
                    pbar.update(1) 
                    sleep(0.03) 
            answer = True 

            if answer: 
                if platform.system() == 'Windows': 
                    os.system(f'start {os.path.join(self.new_path, self.file_name)}')
                elif platform.system() == 'Darwin': 
                    subprocess.run(['open', os.path.join(self.new_path, self.file_name)])
                elif platform.system() == 'Linux': 
                    print('Wow you special')
                    subprocess.run(['xdg-open', os.path.join(self.new_path, self.file_name)])

        except FileNotFoundError as e: 
            print(f'Error has occured: {e}')
        finally:
            print('Completed')

# data = pd.DataFrame({
#     'col1': ['hello this', 'is just'],
#     'col2': ['random strings', 'in four'],
#     'col3': ['random columns', 'for testing'],
#     'col4': ['to see if the columns', 'will adjust dynamically with this method']
# })
# logger = DataExport(data, file_name='my_logs.xlsx', directory='my_logs')
