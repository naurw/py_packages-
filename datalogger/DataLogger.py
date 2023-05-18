import tkinter as tk
from tkinter import messagebox
import os
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import workbook, Workbook 
from openpyxl.styles import PatternFill
from time import sleep 
from contextlib import contextmanager
import mimetypes 

from tqdm import tqdm 
import platform 
import subprocess

# os.getcwd()
# import sys 
# sys.path.append(os.path.join(os.getcwd(), 'Desktop/py_packages-/datalogger/gui_add_ons'))
# import pbar 
# from .gui_add_ons.pbar import ProgressBar 

workbook.WB_VIEW_NORMAL = 'Sheet View'

class DataLogger:
    def __init__(self, dataframe, file_name='v3_speed_test_logs.xlsx', directory='v3_logs'):
        self.df = dataframe
        self.current_path = os.getcwd()
        self.file_name = file_name
        self.directory = directory
        self.new_path = ''
        self.gui = tk.Tk() 
        self.gui.withdraw() 
        self.max_col_width = None
        self.destroyed = False 

        self.check_directory()
        self.write_to_excel()

        if self.gui.winfo_exists():
            self.gui.event_generate("<<ThemeChanged>>") # addresses error when application is closed or destroyed before event being generated is processed. if window has already been destroyed, then event will not be genereated 

    def check_directory(self):
        logs_dir = os.path.join(self.current_path, self.directory)
        if not os.path.exists(logs_dir):
            print(f'File Path: {logs_dir} does not exist... creating directory now')
            message = f'File Path: {logs_dir} does not exist... creating directory now'
            messagebox.showinfo(message=message)
            os.mkdir(logs_dir)
        else:
            print(f'File Path: {logs_dir} exists... changing directory')
            message = f'File Path: {logs_dir} exists... changing directory'
            messagebox.showinfo(message=message)
        os.chdir(logs_dir)
        self.new_path = os.getcwd()
        print(self.new_path)
    
    def adjust_col_width(self, sheet): 
        if self.max_col_width is None:  
            self.set_max_col_width()

        for col in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = self.get_max_cell_length(sheet[column_letter])
            adjusted_width = self.max_col_width if self.max_col_width is not None and (self.max_col_width >= max_length or self.max_col_width <= max_length + 5) else min(self.max_col_width, max_length +1)
            sheet.column_dimensions[column_letter].width = adjusted_width 

    def set_max_col_width(self): 
        try: 
            self.max_col_width = int(input(f'Enter the maximum desired column width (1-50): '))
            if self.max_col_width > 50 or self.max_col_width < 1: 
                print(f'Invalid input. Using default max value of 50.')
                self.max_col_width = 50

        except ValueError: 
            print('Invalid input. Using default max value of columns.')
            self.max_col_width = self.get_max_cell_length

    def get_max_cell_length(self, column_cells):
        max_length = 0 
        for cell in column_cells: 
            try: 
                cell_value_length = len(str(cell.value))
                if cell_value_length > max_length: 
                    max_length = cell_value_length
                    print(max_length)

            except Exception as e: 
                print(f'Error occured: {e}')

        return max_length 


###############################################################
### Simpler method for computing column widths using pixels ###
###############################################################
# from openpyxl.utils import column_width_from_pixel
#     def adjust_col_width_pixel(self, sheet): 
#         from openpyxl.utils import column_width_from_pixel
#         for col in sheet.columns:
#             max_length = max(len(str(cell.value)) for cell in col)
#             adjusted_width = column_width_from_pixel(max_length * 7)  # assuming default font size of 11pt
#             col[0].column_dimensions.width = adjusted_width

    def write_to_excel(self):
        try:
            file_path = os.path.join(self.new_path, self.file_name)
            if not os.path.exists(file_path):
                message = f'{file_path} does not exist... creating file'
                messagebox.showinfo(title='Warning', message=message)

                if isinstance(self.df, pd.DataFrame) and not isinstance(self.df,list): # checks if input is a single DataFrame object 
                    with pd.ExcelWriter(file_path, engine = 'openpyxl', mode = 'w') as writer: 
                        self.df.to_excel(writer, sheet_name = 'Sheet1', index=False)
                        sheet = writer.book['Sheet1']
                        self.adjust_col_width(sheet)
                        
                elif isinstance(self.df, list) and all(isinstance(df, pd.DataFrame) for df in self.df): # checks if input is a list that ONLY contains dataframes 
                    with pd.ExcelWriter(file_path, engine = 'openpyxl', mode = 'w') as writer: 
                        for i, df in enumerate(self.df): # iterate through the elements and creating an index for naming
                            sheet_name = df.name if hasattr(df, 'name') else f'Sheet{i+1}' # check for whether if dataframe(s) have attr 'name' and if it does wil be used 
                            df.to_excel(writer, sheet_name = sheet_name, index = False) # either use existing name as sheet or use default with numerical suffix
                            sheet = writer.book[sheet_name]
                            self.adjust_col_width(sheet)

                else: 
                    print('Var used is neither a dataframe nor a list of dataframes. Retry.')
                    raise ValueError
                
            else:
                message = f'{file_path} has been found... opening and appending'
                messagebox.showinfo(title='File Found', message=message)

                if isinstance(self.df, pd.DataFrame) and not isinstance(self.df,list): 
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        self.df.to_excel(writer, sheet_name='Sheet1', index=False)
                        sheet = writer.book['Sheet1']
                        self.adjust_col_width(sheet)
                
                elif isinstance(self.df, list) and all(isinstance(df, pd.DataFrame) for df in self.df): # checks if input is a list that ONLY contains dataframes 
                    with pd.ExcelWriter(file_path, engine = 'openpyxl', mode = 'w') as writer: 
                        for i, df in enumerate(self.df): # iterate through the elements and creating an index for naming
                            sheet_name = df.name if hasattr(df, 'name') else f'Sheet{i+1}' # check for whether if dataframe(s) have attr 'name' and if it does wil be used 
                            df.to_excel(writer, sheet_name = sheet_name, index = False) # either use existing name as sheet or use default with numerical suffix 
                            sheet = writer.book[sheet_name]
                            self.adjust_col_width(sheet)

                else: 
                    print('Var used is neither a dataframe nor a list of dataframes. Retry.')
                    raise ValueError 
                
            
            os.chdir(os.pardir)

            self.open_file() 
        except Exception as e:
            message = f'Error occurred:\n{e}\n\nRetrying with alternative method'
            messagebox.showinfo(title='Warning', message=message)
            sleep(3)

            if isinstance(self.df, pd.DataFrame) and not isinstance(self.df, list): 
                with pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a', if_sheet_exists='replace', engine_kwargs = {'options' : {'sheet_state' : 'visible'}}) as writer:
                    try: 
                        if mimetypes.guess_type(self.file_name)[0] in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                            print('File is the formats: vnd.ms-excel or vnd.openxml formats')
                            existing_df = pd.read_excel(self.file_name)
                        else: 
                            print('File is not the in the accepted pd.read_excel formats... using alternative method') 
                            existing_file = pd.ExcelFile(self.file_name)
                            existing_df = existing_file.parse(sheet_name = 'Sheet1')
                    except Exception as e: 
                        print(f'Error occured: {e}')
        
                    book = load_workbook(self.file_name)
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in writer.book.worksheets} # get all the sheets 

                    existing_columns = set() # extract existing column names form the first sheet 
                    for sheet in book.worksheets:
                        if sheet.title == 'Sheet1':
                            existing_df_sheet = pd.read_excel(self.file_name, sheet_name=sheet.title)
                            existing_columns.update(existing_df_sheet.columns)


                    new_columns = set(self.df.columns) # check if new data has the same column names as the existing data 
                    if existing_columns == new_columns:
                        new_df = pd.concat([existing_df_sheet, self.df], axis=0, ignore_index=True)
                        last_row = len(existing_df_sheet)
                        new_df.style.apply(lambda x: ['background-color: #C4D9F2' if x.name >= last_row else ''], axis=1) # change color for only newly appended data and default white if data is old 
                        new_df.to_excel(writer, sheet_name='Sheet1', index=False)
                        sheet = book['Sheet1']
                        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, values_only=True), start=2): # iterates over the rows of the worksheet starting from a set starting row. returns iterable of Cell objects for each row. Ensures we get the cell values instead of ALL of the cells' objects. Enumerate to get both the row index and the row values during iteration. Start =2 sets the starting value of the row index
                            for col_idx, value in enumerate(row, 1):
                                if row_idx > last_row + 1: 
                                    cell = sheet.cell(row = row_idx, column = col_idx)
                                    cell.fill = PatternFill(start_color='C4D9F2', end_color='C4D9F2', fill_type='solid')
                                
                        self.adjust_col_width(sheet)
                    else:
                        self.df.to_excel(writer, sheet_name='Sheet1', index=False)
                        sheet = book['Sheet1']
                        self.adjust_col_width(sheet)
                                

            elif isinstance(self.df, list) and all(isinstance(df, pd.DataFrame) for df in self.df):
                with pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a', if_sheet_exists='replace', engine_kwargs = {'options' : {'sheet_state' : 'visible'}}) as writer:
                        try: 
                            if mimetypes.guess_type(self.file_name)[0] in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                                print('File is the formats: vnd.ms-excel or vnd.openxml formats')
                                existing_df = pd.read_excel(self.file_name)
                            else: 
                                print('File is not the in the accepted pd.read_excel formats... using alt method') 
                                existing_file = pd.ExcelFile(self.file_name)
                                existing_df = existing_file.parse(sheet_name = 'Sheet1')
                        except Exception as e: 
                            print(f'Error occured: {e}')

                        new_data = []
                        existing_columns = set() # Extract column names from the existing data into a set (different from a list because unique values only) 
                        for sheet_name, df in existing_df.items():
                            existing_columns.update(df.columns)

                        for df in self.df:
                            new_columns = set(df.columns) # Extract column names from the new dataframes into a set

                            if existing_columns.intersection(new_columns): # If there are any overlapping columns, concatenate the new data with the existing data and add to the list of new data to be written
                                existing_df_sheet = existing_df.get(df.name, pd.DataFrame()) # if there is a dataframe name then it takes that. if there isn't then default value 
                                last_row = len(existing_df_sheet)
                                new_df = pd.concat([existing_df_sheet, df], axis=0, ignore_index=True)
                                new_df.style.apply(lambda x: ['background-color: #C4D9F2' if x.name >= last_row else ''], axis=1) # temporary function for creating background colors 
                                new_data.append(new_df)
                                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, values_only=True), start=2): # iterates over the rows of the worksheet starting from a set starting row. returns iterable of Cell objects for each row. Ensures we get the cell values instead of ALL of the cells' objects. Enumerate to get both the row index and the row values during iteration. Start =2 sets the starting value of the row index
                                    for col_idx, value in enumerate(row, 1):
                                        if row_idx > last_row + 1: 
                                            cell = sheet.cell(row = row_idx, column = col_idx)
                                            cell.fill = PatternFill(start_color='C4D9F2', end_color='C4D9F2', fill_type='solid')
                            else:
                                new_data.append(df) # If there are no overlapping columns, it treats it as a new dataframe and add it to the list of dataframes 
                            
                        for df in new_data:
                            df.to_excel(writer, sheet_name=df.name, index=False)
                            sheet = writer.sheets[df.name]
                            self.adjust_col_width(sheet)
            else: 
                print('Var used is neither a dataframe nor a list of dataframes. Retry.')
                raise ValueError 
            
            os.chdir(os.pardir)

            self.open_file() 

    def open_file(self): 
        try: 
            # for i in tqdm(range(100), desc='Opening file', unit='%', ncols=70): # progress bar length set to 70 chars wide 
            #     sleep(0.03) # animation set at 0.03

            with tqdm(total = 100, desc = 'Opening File', unit = '%', ncols = 70) as pbar: 
                for i in range(100): 
                    pbar.update(1) 
                    sleep(0.03) 

            answer = messagebox.askyesno(title = 'File', message = 'Do you want to open file?')
            if answer: 
                if platform.system() == 'Windows': 
                    os.system(f'start {os.path.join(self.new_path, self.file_name)}')
                elif platform.system() == 'Darwin': 
                    subprocess.run(['open', os.path.join(self.new_path, self.file_name)])
                elif platform.system() == 'Linux': 
                    print('Wow you special')
                    subprocess.run(['xdg-open', os.path.join(self.new_path, self.file_name)])

        except FileNotFoundError as e: 
            print(f'Error has occured: {e}')
        finally:
            if not self.destroyed: # add a flag for handling if .destroyed() 
                self.gui.destroy() 
                self.destroyed = True 
            
            print(self.destroyed)
            print('Completed')


        self.gui.mainloop() # potential infinite looping / leaked semaphore(s) 
        
        
## Testing for singular dataframes 
data = pd.DataFrame({
    'col1': ['hello this', 'is just'],
    'col2': ['random strings', 'in four'],
    'col3': ['random columns', 'for testing'],
    'col4': ['to see if the columns', 'will adjust dynamically with this method']
})
logger = DataLogger(data, file_name='my_logs.xlsx', directory='my_logs')


## Testing for multiple dataframes 
import numpy as np 
np.random.seed(42)

arr = np.random.rand(2,4)
df1 = pd.DataFrame(arr, columns= ['col1', 'col2', 'col3', 'col4'])
df1.name = 'np array 1'
arr2 = np.random.rand(2,4) * 100
df2 = pd.DataFrame(arr2, columns = ['col1', 'col2', 'col3', 'col4'])

df_list = [data, df1, df2]
logger_mult = DataLogger(df_list, file_name='my_logs.xlsx', directory='my_logs')
logger_mult

