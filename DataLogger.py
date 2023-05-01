import tkinter as tk
from tkinter import messagebox
from tkinter.messagebox import askyesno
import os
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import workbook, Workbook 
from time import sleep 
from contextlib import contextmanager
import mimetypes 
import platform 
import subprocess 
from tqdm import tqdm 
from pbar import ProgressBar 

# import sys 
# sys.path.append('/Users/William/Desktop/py_packages-')
# from pbar import ProgressBar 

workbook.WB_VIEW_NORMAL = 'Sheet View'

class DataLogger:
    def __init__(self, dataframe, file_name='v3_speed_test_logs.xlsx', directory='v3_logs'):
        self.df = dataframe
        self.current_path = os.getcwd()
        self.file_name = file_name
        self.directory = directory
        self.new_path = ''
        self.gui = tk.Tk() 
        self.gui.withdraw() 

        self.check_directory()
        self.write_to_excel()

    def check_directory(self):
        logs_dir = os.path.join(self.current_path, self.directory)
        if not os.path.exists(logs_dir):
            print(f'File Path: {logs_dir} does not exist... creating directory now')
            message = f'File Path: {logs_dir} does not exist... creating directory now'
            messagebox.showinfo(message=message)
            os.mkdir(logs_dir)
        else:
            print(f'File Path: {logs_dir} exists... changing directory')
            message = f'File Path: {logs_dir} exists... changing directory'
            messagebox.showinfo(message=message)
        os.chdir(logs_dir)
        self.new_path = os.getcwd()
        print(self.new_path)
    
    def adjust_col_width(self, sheet): 
        try: 
            max_col_width = int(input('Enter the maximum desired column width: '))
            if max_col_width > 50:
                print(f'Value too high. Using default max value of 50.')
                max_col_width = 50
                for col in range(1, sheet.max_column + 1): 
                    column_letter = get_column_letter(col)
                    max_length = 0 
                    for cell in sheet[column_letter]:
                        try: 
                            if len(str(cell.value)) > max_length: 
                                max_length = len(str(cell.value)) 
                        except Exception as e: 
                            print(f'Error Occured: {e}')

                    adjusted_width = min(max_col_width, max_length + 1) # allows for user adjustments while selecting the mininum possible length necessary to keep readability 

                    sheet.column_dimensions[column_letter].width = adjusted_width
        except ValueError: 
            print('Invalid input. Using default max value of column_length.')
            for col in range(1, sheet.max_column + 1): 
                column_letter = get_column_letter(col)
                max_length = 0 
                for cell in sheet[column_letter]:
                    try: 
                        if len(str(cell.value)) > max_length: 
                            max_length = len(str(cell.value)) 
                    except Exception as e: 
                        print(f'Error Occured: {e}')
                adjusted_width = (max_length + 1) 
                sheet.column_dimensions[column_letter].width = adjusted_width
        # for col in range(1, sheet.max_column + 1): 
        #     column_letter = get_column_letter(col)
        #     max_length = 0 
        #     for cell in sheet[column_letter]:
        #         try: 
        #             if len(str(cell.value)) > max_length: 
        #                 max_length = len(str(cell.value)) 
        #         except Exception as e: 
        #             print(f'Error Occured: {e}')

        #     adjusted_width = min(max_col_width, max_length + 1) # allows for user adjustments while selecting the mininum possible length necessary to keep readability 

        #     sheet.column_dimensions[column_letter].width = adjusted_width

### Simpler method for adjust_col_width() 
# from openpyxl.utils import column_width_from_pixel
#     def adjust_col_width_pixel(self, sheet): 
#         from openpyxl.utils import column_width_from_pixel
#         for col in sheet.columns:
#             max_length = max(len(str(cell.value)) for cell in col)
#             adjusted_width = column_width_from_pixel(max_length * 7)  # assuming default font size of 11pt
#             col[0].column_dimension.width = adjusted_width

    def write_to_excel(self):
        try:
            file_path = os.path.join(self.new_path, self.file_name)
            if not os.path.exists(file_path):
                message = f'{file_path} does not exist... creating file'
                messagebox.showinfo(title='Warning', message=message)
                with pd.ExcelWriter(file_path, engine = 'openpyxl', mode = 'w') as writer: 
                    self.df.to_excel(writer, sheet_name = 'Sheet1', index=False)

                    # adjust_col_width() function added to the very first file in case errors persist
                    wb = writer.book
                    for sheet in wb.worksheets: 
                        for col in range(1, sheet.max_column + 1): 
                            column_letter = get_column_letter(col)
                            max_length = 0 
                            for cell in sheet[column_letter]:
                                try: 
                                    if len(str(cell.value)) > max_length: 
                                        max_length = len(str(cell.value)) 
                                except Exception as e: 
                                    print(f'Error Occured: {e}')
                            adjusted_width = (max_length + 1) 
                            sheet.column_dimensions[column_letter].width = adjusted_width 
                
            else:
                message = f'{file_path} has been found... opening and appending'
                messagebox.showinfo(title='File Found', message=message)

                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    self.df.to_excel(writer, sheet_name='Sheet1', index=False)
                    for sheet in writer.book.worksheets: 
                        self.adjust_col_width(sheet)
        
            os.chdir(os.pardir)

            self.open_file() 
        except Exception as e:
            message = f'Error occurred:\n{e}\n\nRetrying with alternative method'
            messagebox.showinfo(title='Warning', message=message)
            sleep(3)

            with pd.ExcelWriter(self.file_name, engine='openpyxl', mode='a', if_sheet_exists='replace', engine_kwargs = {'options' : {'sheet_state' : 'visible'}}) as writer:
                try: 
                    if mimetypes.guess_type(self.file_name)[0] in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                        print('File is the formats: vnd.ms-excel or vnd.openxml formats')
                        existing_df = pd.read_excel(self.file_name)
                    else: 
                        print('File is not the in the accepted pd.read_excel formats... using alt method') 
                        existing_file = pd.ExcelFile(self.file_name)
                        existing_df = existing_file.parse(sheet_name = 'Sheet1')
                except Exception as e: 
                    print(f'Error occured: {e}')

                # messagebox.showinfo(title = 'Previous Data', message = f'\n\n{existing_df.tail(5)}')
                book = load_workbook(self.file_name)
                writer.book = book
                # active_sheet = book.active 
                # messagebox.showinfo(message = f'{self.file_name} is currently active\n\nActive Sheet: {active_sheet}\nSheet Title:{active_sheet.title}')
                writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
                new_df = pd.concat([existing_df, self.df], axis=0, ignore_index=True)
                # messagebox.showinfo(title = 'Updated Data', message = f'\n\n{new_df.tail(5)}')

                sheet_names = book.sheetnames 
                for sheet_name in sheet_names: 
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    sheet = book[sheet_name]
                    self.adjust_col_width(sheet) 

            os.chdir(os.pardir)

            self.open_file() 

    def open_file(self): 
        try: 
            # for i in tqdm(range(100), desc='Opening file', unit='%', ncols=70): # progress bar length set to 70 chars wide 
            #     sleep(0.03) # animation set at 0.03

            # with tqdm(total = 100, desc = 'Opening File', unit = '%', ncols = 70) as pbar: 
            #     for i in range(100): 
            #         pbar.update(1) 
            #         sleep(0.03) 



            # answer = messagebox.askyesno(title = 'File', message = 'Do you want to open file?')
            # if answer: 
            #     if platform.system() == 'Windows': 
            #         os.system(f'start {os.path.join(self.new_path, self.file_name)}')
            #     elif platform.system() == 'Darwin': 
            #         subprocess.run(['open', os.path.join(self.new_path, self.file_name)])
            #     elif platform.system() == 'Linux': 
            #         print('Wow you special')
            #         subprocess.run(['xdg-open', os.path.join(self.new_path, self.file_name)])
            

            pbar = ProgressBar(file_name = self.file_name, directory = self.directory)
            pbar.find_folder() 
            pbar.open_file() 
        except FileNotFoundError as e: 
            print(f'Error has occured: {e}')
        finally: 
            print('Completed')
            self.gui.destroy()

        self.gui.mainloop()
        
        

data = pd.DataFrame({
    'col1': ['hello this', 'is just'],
    'col2': ['random strings', 'in four'],
    'col3': ['random columns', 'for testing'],
    'col4': ['to see if the columns', 'will adjust dynamically with this method']
})
logger = DataLogger(data, file_name='my_logs.xlsx', directory='my_logs')
logger







### Sandbox for DataLoggerMulti: 
# import pandas as pd

# # List of dataframes
# dfs = [df1, df2, df3]

# # Create a Pandas Excel writer using XlsxWriter as the engine.
# writer = pd.ExcelWriter('output.xlsx', engine='xlsxwriter')

# # Loop through the list of dataframes and write each to a different sheet in the workbook
# for i, df in enumerate(dfs):
#     sheet_name = 'Sheet{}'.format(i+1)
#     df.to_excel(writer, sheet_name=sheet_name, index=False)

# # Close the Pandas Excel writer and output the Excel file.
# writer.save()

### v3.1 
class DataLoggerMulti: 
    def __init__(self, file_name): 
        self.file_name = file_name
        self.book = None
        self.sheet_names = [] 
    
    def __enter__(self): 
        self.writer = pd.ExcelWriter(self.file_name, engine = 'openpyxl')
        return self 

    def write_dataframes_to_excel(self, dfs): 
        if os.path.exists(self.file_name): 
            self.book = load_workbook(self.file_name)
        else: 
            Workbook() 
            # self.writer.save() 
            print('File does not exist') 
        # self.book = load_workbook(self.file_name) if os.path.exists(self.file_name) else Workbook()
        self.writer.book = self.book 
        self.sheet_names = self.book.sheetnames  

        for df in dfs: 
            sheet_name = df.name if hasattr(df, 'name') else 'Sheet' # <-- Check for whether if dataframe(s) have attr 'name' 
            if sheet_name in self.sheet_names: 
                sheet = self.book[sheet_name]
                self.book.remove(sheet) 

            df.to_excel(self.writer, sheet_name=sheet_name, index = False) 
            self.sheet_names.append(sheet_name) 

            self.writer.save() 

    def close_writer(self): 
        self.writer.close()

    def __exit__(self, exc_type, exc_val, exc_tb): 
        self.writer.close()  

# with DataLoggerMulti('my_file.xlsx') as writer: 
#     writer.write_dataframes_to_excel([df1, df2, df3])



